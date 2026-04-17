#!/usr/bin/env python3
"""Turn EXECUTION_PLAN_V2.xlsx's Daily Schedule sheet into GitHub issues.

Uses the `gh` CLI (not the REST API directly) so it picks up the user's
existing auth / repo context automatically. Creates one issue per scheduled
day, labels it with priority + category + phase, and places it in the
appropriate phase milestone.

Usage:
    # Dry run (default) — prints what would be created
    python scripts/create_github_issues.py

    # Actually create the issues on the current repo
    python scripts/create_github_issues.py --go

    # Only create days 1-20 (for incremental sync)
    python scripts/create_github_issues.py --go --from 1 --to 20

    # Target a different repo
    python scripts/create_github_issues.py --go --repo owner/name

Requires:
    - `gh` CLI installed and authenticated (`gh auth login`)
    - openpyxl (already in dev deps)
"""
from __future__ import annotations

import argparse
import shutil
import subprocess
import sys
from pathlib import Path

from openpyxl import load_workbook

PLAN_XLSX = Path(__file__).resolve().parents[1] / "EXECUTION_PLAN_V2.xlsx"
PHASE_LABELS = {
    "1. Foundation":       "phase/1-foundation",
    "2. First Channel":    "phase/2-first-channel",
    "3. Production":       "phase/3-production",
    "4. Portfolio":        "phase/4-portfolio",
    "5. Scale":            "phase/5-scale",
    "6. Diversify":        "phase/6-diversify",
    "7. Target Hit":       "phase/7-target",
}
CATEGORY_LABELS = {
    "Infra":         "cat/ops",
    "Content":       "cat/content",
    "Marketing":     "cat/marketing",
    "Optimization":  "cat/optimization",
    "Analytics":     "cat/analytics",
    "Ops":           "cat/ops",
    "Rest":          "cat/ops",
}


def _check_gh():
    if not shutil.which("gh"):
        sys.stderr.write(
            "gh CLI not found. Install from https://cli.github.com/ and run `gh auth login`.\n"
        )
        sys.exit(2)


def _existing_milestones(repo: str | None) -> set[str]:
    cmd = ["gh", "api", "repos/:owner/:repo/milestones", "--jq", ".[].title"]
    if repo:
        cmd = ["gh", "api", f"repos/{repo}/milestones", "--jq", ".[].title"]
    try:
        out = subprocess.run(cmd, capture_output=True, text=True, check=True).stdout
    except subprocess.CalledProcessError:
        return set()
    return {line.strip() for line in out.splitlines() if line.strip()}


def _ensure_milestone(title: str, description: str, repo: str | None, go: bool) -> None:
    existing = _existing_milestones(repo)
    if title in existing:
        return
    print(f"  + creating milestone: {title}")
    if not go:
        return
    cmd = ["gh", "api", "repos/:owner/:repo/milestones",
           "-f", f"title={title}", "-f", f"description={description}"]
    if repo:
        cmd[2] = f"repos/{repo}/milestones"
    subprocess.run(cmd, check=True, capture_output=True)


def _ensure_label(name: str, color: str, description: str, repo: str | None, go: bool) -> None:
    cmd = ["gh", "label", "create", name, "--color", color, "--description", description]
    if repo:
        cmd.extend(["--repo", repo])
    if not go:
        print(f"  + would create label {name}")
        return
    subprocess.run(cmd, capture_output=True, text=True)
    # Ignore exit code — 'already exists' is fine


def _create_issue(title: str, body: str, labels: list[str], milestone: str,
                  repo: str | None, go: bool) -> None:
    print(f"  + issue: {title}  [{', '.join(labels)}]")
    if not go:
        return
    cmd = ["gh", "issue", "create",
           "--title", title,
           "--body", body,
           "--milestone", milestone]
    for lbl in labels:
        cmd.extend(["--label", lbl])
    if repo:
        cmd.extend(["--repo", repo])
    subprocess.run(cmd, check=True, capture_output=True, text=True)


def main() -> int:
    p = argparse.ArgumentParser(description=__doc__)
    p.add_argument("--go", action="store_true",
                   help="Actually create issues (default: dry run)")
    p.add_argument("--repo", default=None,
                   help="owner/name of target repo (default: current)")
    p.add_argument("--from", dest="from_day", type=int, default=1)
    p.add_argument("--to", dest="to_day", type=int, default=180)
    p.add_argument("--plan", default=str(PLAN_XLSX))
    args = p.parse_args()

    if args.go:
        _check_gh()

    plan_path = Path(args.plan)
    if not plan_path.exists():
        sys.stderr.write(f"Plan not found: {plan_path}\n")
        return 2

    wb = load_workbook(plan_path, read_only=True, data_only=True)
    if "Daily Schedule" not in wb.sheetnames:
        sys.stderr.write("Sheet 'Daily Schedule' not found in plan.\n")
        return 2

    ws = wb["Daily Schedule"]
    rows = list(ws.iter_rows(min_row=2, values_only=True))

    # ── Ensure all phase milestones + category labels exist ──────────
    print("── Ensuring milestones ──")
    for phase, _ in PHASE_LABELS.items():
        _ensure_milestone(f"Phase {phase}", f"180-day plan phase: {phase}",
                          args.repo, args.go)

    print("\n── Ensuring phase labels ──")
    for _, lbl in PHASE_LABELS.items():
        _ensure_label(lbl, "5319e7", "Phase label from EXECUTION_PLAN_V2",
                      args.repo, args.go)

    # ── Create one issue per scheduled day ───────────────────────────
    print(f"\n── Creating issues for days {args.from_day}..{args.to_day} "
          f"({'GO' if args.go else 'DRY RUN'}) ──")
    n = 0
    for row in rows:
        if not row or row[0] is None:
            continue
        day, d_date, week, month, phase, task, category, priority, hrs, deps, _ = row
        if day < args.from_day or day > args.to_day:
            continue

        title = f"[Day {day}] {task}"
        if len(title) > 256:
            title = title[:253] + "..."
        body = (
            f"**Scheduled for:** {d_date}  \n"
            f"**Day:** {day}  •  **Week:** {week}  •  **Month:** {month}  \n"
            f"**Phase:** {phase}  \n"
            f"**Category:** {category}  \n"
            f"**Priority:** {priority}  \n"
            f"**Estimated hours:** {hrs}  \n"
            f"**Dependencies:** {deps or '_(none)_'}  \n\n"
            f"---\n\n"
            f"Task from `EXECUTION_PLAN_V2.xlsx` → `Daily Schedule` sheet, row {day + 1}.\n"
        )
        labels = [
            priority,
            CATEGORY_LABELS.get(category, "cat/ops"),
            PHASE_LABELS.get(phase, "task"),
            "task",
        ]
        milestone = f"Phase {phase}"
        _create_issue(title, body, labels, milestone, args.repo, args.go)
        n += 1

    print(f"\nDone. {n} issue(s) {'created' if args.go else 'would be created'}.")
    return 0


if __name__ == "__main__":
    sys.exit(main())
